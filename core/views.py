from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
from math import ceil
from django.db import IntegrityError, transaction
from datetime import date, timedelta, datetime
from io import BytesIO, StringIO
import csv
from django.db.models import Q, Count
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.utils.encoding import smart_str
from .models import Titulo, Ejemplar, Prestamo, Categoria, ensure_categoria_otros
from .forms import LibroForm, PrestamoForm


# -------------------------------------------------------------------
# Datos demo (los dejé como los tenías, no se usan en las vistas DB)
# -------------------------------------------------------------------
LIBROS = [
    {"titulo":"Cien años de soledad","autor":"García Márquez","categoria":"Psicologia","estado":"Disponible"},
    {"titulo":"Breves respuestas a grandes preguntas","autor":"Stephen Hawking","categoria":"Ciencia","estado":"Disponible"},
    {"titulo":"Fundamentos de la Educacion Social","autor":"Jorge Lopez","categoria":"Social","estado":"Prestado"},
    {"titulo":"It","autor":"Stephen King","categoria":"Ciencia","estado":"Disponible"},
    {"titulo":"Odisea","autor":"Homero","categoria":"Social","estado":"Prestado"},
    {"titulo":"Derecho societario","autor":"Ricardo Augusto Lopez","categoria":"Social","estado":"Disponible"},
    {"titulo":"El lenguaje de las emociones","autor":"Juan Perez","categoria":"Psicologia","estado":"Prestado"},
    {"titulo":"El poder de las palabras","autor":"Mariano Sigman","categoria":"Social","estado":"Disponible"},
    {"titulo":"El Alquimista","autor":"Paulo Coelho","categoria":"Ciencia","estado":"Disponible"},
    {"titulo":"El Extranjero","autor":"Albert Camus","categoria":"Social","estado":"Disponible"},
    {"titulo":"Divina Comedia","autor":"Dante Alighieri","categoria":"Psicologia","estado":"Disponible"},
    {"titulo":"Crimen y castigo","autor":"Fiódor Dostoyevski","categoria":"Psicologia","estado":"Disponible"},
    {"titulo":"El origen de las especies","autor":"Charles Darwin","categoria":"Ciencia","estado":"Disponible"},
    {"titulo":"Sapiens: De animales a dioses","autor":"Yuval Noah Harari","categoria":"Social","estado":"Prestado"},
    {"titulo":"El hombre en busca de sentido","autor":"Viktor Frankl","categoria":"Psicologia","estado":"Disponible"},
    {"titulo":"Breve historia del tiempo","autor":"Stephen Hawking","categoria":"Ciencia","estado":"Prestado"},
    {"titulo":"Los miserables","autor":"Victor Hugo","categoria":"Social","estado":"Disponible"},
    {"titulo":"El arte de la guerra","autor":"Sun Tzu","categoria":"Social","estado":"Disponible"},
    {"titulo":"La interpretación de los sueños","autor":"Sigmund Freud","categoria":"Psicologia","estado":"Prestado"},
    {"titulo":"Cosmos","autor":"Carl Sagan","categoria":"Ciencia","estado":"Disponible"},
    {"titulo":"1984","autor":"George Orwell","categoria":"Social","estado":"Disponible"},
    {"titulo":"El príncipe","autor":"Nicolás Maquiavelo","categoria":"Social","estado":"Prestado"},
    {"titulo":"Pensar rápido, pensar despacio","autor":"Daniel Kahneman","categoria":"Psicologia","estado":"Disponible"}
]
PRESTAMOS = [
    {"alumno":"Juan Pérez","libro":"Fundamentos de la Educacion Social","fecha_prestamo":"2025-09-02","vence":"2025-09-23","devuelto":True},
    {"alumno":"Lucía Fernández","libro":"La interpretación de los sueños","fecha_prestamo":"2025-09-13","vence":"2025-09-19","devuelto":False},
    {"alumno":"Diego Martínez","libro":"El príncipe","fecha_prestamo":"2025-09-14","vence":"2025-09-27","devuelto":False},
    {"alumno":"Micaela Cortina","libro":"El lenguaje de las emociones","fecha_prestamo":"2025-09-03","vence":"2025-09-17","devuelto":False},
    {"alumno":"Roberto García","libro":"Odisea","fecha_prestamo":"2025-08-28","vence":"2025-09-13","devuelto":False},
    {"alumno":"Nicolás Lopez","libro":"Sapiens: De animales a dioses","fecha_prestamo":"2025-09-05","vence":"2025-09-28","devuelto":False},
    {"alumno":"Pilar Kansas","libro":"Breve historia del tiempo","fecha_prestamo":"2025-09-01","vence":"2025-09-11","devuelto":False}
]


# -------------------------------------------------------------------
# Dashboard
# -------------------------------------------------------------------
def dashboard(request):
    hoy = date.today()

    stock_total = Ejemplar.objects.count()
    prestados = Ejemplar.objects.filter(estado__iexact="Prestado").count()
    disponibles = Ejemplar.objects.filter(estado__iexact="Disponible").count()
    ocupacion = round((prestados / stock_total) * 100, 1) if stock_total else 0

    vence_7d_qs = (
        Prestamo.objects
        .select_related("ejemplar__titulo")
        .filter(estado__in=["ACTIVO","RENOVADO"], vence__gt=hoy, vence__lte=hoy + timedelta(days=7))
        .values("alumno_nombre", "ejemplar__titulo__titulo", "vence")
        .order_by("vence")
    )
    atrasados_qs = (
        Prestamo.objects
        .select_related("ejemplar__titulo")
        .filter(estado__in=["ACTIVO","RENOVADO"], vence__lt=hoy)
        .values("alumno_nombre", "ejemplar__titulo__titulo", "vence")
        .order_by("vence")
    )

    vence_7d = [{"alumno": d["alumno_nombre"], "libro": d["ejemplar__titulo__titulo"], "vence": d["vence"]} for d in vence_7d_qs]
    atrasados = [{"alumno": d["alumno_nombre"], "libro": d["ejemplar__titulo__titulo"], "vence": d["vence"]} for d in atrasados_qs]

    return render(request, "dashboard.html", {
        "total": stock_total,
        "prestados": prestados,
        "disponibles": disponibles,
        "ocupacion": ocupacion,
        "vence_7d": vence_7d,
        "atrasados": atrasados,
    })


# -------------------------------------------------------------------
# Listado de libros (con conteo de stock DISPONIBLE/PRESTADO)
# -------------------------------------------------------------------
def libros_list(req):
    q = (req.GET.get("q") or "").strip()
    cat = (req.GET.get("cat") or "").strip()
    page = int(req.GET.get("page", 1))
    per_page = int(req.GET.get("per_page", 8))
    per_page_choice = [5, 10, 20]

    qs = (
        Titulo.objects
        .annotate(
            disponibles=Count("ejemplares", filter=Q(ejemplares__estado="DISPONIBLE")),
            prestados=Count("ejemplares",   filter=Q(ejemplares__estado="PRESTADO")),
        )
        .prefetch_related("categorias")
        .order_by("titulo")
    )

    if q:
        qs = qs.filter(Q(titulo__icontains=q) | Q(autor__icontains=q))
    if cat:
        qs = qs.filter(categorias__nombre__iexact=cat)

    data = []
    for t in qs:
        cats = ", ".join(t.categorias.values_list("nombre", flat=True))
        data.append({
            "id":t.id,
            "titulo": t.titulo,
            "autor": t.autor,
            "categorias": cats or "—",
            "disponibles": t.disponibles,
            "prestados": t.prestados,
        })

    total = len(data)
    pages = max(1, ceil(total / per_page)) if total else 1
    page = max(1, min(page, pages))
    start, end = (page - 1) * per_page, (page - 1) * per_page + per_page
    page_items = data[start:end]

    cats = list(Categoria.objects.order_by("nombre").values_list("nombre", flat=True))

    return render(req, "libros_list.html", {
        "libros": page_items,
        "cats": cats,
        "q": q, "cat": cat,
        "page": page, "pages": pages, "total": total,
        "per_page": per_page, "per_page_choice": per_page_choice,
        "querystring_base": f"q={q}&cat={cat}&per_page={per_page}",
        "page_range": range(1, pages + 1),
    })

def libros_export(request):
    """
    Exporta a Excel (.xlsx) el catálogo: Título, Autor, Categorías, Disponibles, Prestados, Stock total.
    Respeta filtros por ?q=&cat= si vienen en la querystring.
    Si no está disponible openpyxl, cae a CSV automáticamente.
    """
    q = (request.GET.get("q") or "").strip()
    cat = (request.GET.get("cat") or "").strip()

    qs = (
        Titulo.objects
        .annotate(
            disponibles=Count("ejemplares", filter=Q(ejemplares__estado="DISPONIBLE")),
            prestados=Count("ejemplares",   filter=Q(ejemplares__estado="PRESTADO")),
        )
        .prefetch_related("categorias")
        .order_by("titulo")
    )
    if q:
        qs = qs.filter(Q(titulo__icontains=q) | Q(autor__icontains=q))
    if cat:
        qs = qs.filter(categorias__nombre__iexact=cat)

    rows = []
    for t in qs:
        cats = ", ".join(t.categorias.values_list("nombre", flat=True))
        disp = t.disponibles or 0
        pres = t.prestados or 0
        rows.append([
            smart_str(t.titulo),
            smart_str(t.autor),
            smart_str(cats or "—"),
            disp,
            pres,
            disp + pres,  # Stock total
        ])

    filename_base = "catalogo_libros"
    if q or cat:
        filename_base += "_filtrado"
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename_xlsx = f"{filename_base}_{stamp}.xlsx"
    filename_csv  = f"{filename_base}_{stamp}.csv"

    # Intento XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Libros"

        headers = ["Título", "Autor", "Categorías", "Disponibles", "Prestados", "Stock total"]
        ws.append(headers)
        for r in rows:
            ws.append(r)

        # Auto-anchos sencillos
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for r in rows:
                max_len = max(max_len, len(str(r[col_idx-1])) if r[col_idx-1] is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        resp = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename_xlsx}"'
        return resp

    except Exception:
        # Fallback CSV si no hay openpyxl
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow(["Título", "Autor", "Categorías", "Disponibles", "Prestados", "Stock total"])
        writer.writerows(rows)
        resp = HttpResponse(sio.getvalue(), content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename_csv}"'
        return resp


# ---------- LIBROS ----------
@require_http_methods(["GET", "POST"])
def libro_create(request):
    """
    Crea un Titulo + un Ejemplar inicial DISPONIBLE (lo hace el LibroForm.save()).
    Acepta múltiples categorías; si no se elige ninguna, asigna 'Otros'.
    """
    if request.method == "POST":
        form = LibroForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # IMPORTANTE: commit=True (por defecto). Esto evita el error.
                    titulo_obj = form.save()  # crea también el Ejemplar si no existe

                    # Si no se eligió ninguna categoría en el form, agregamos 'Otros'
                    if titulo_obj.categorias.count() == 0:
                        otros = ensure_categoria_otros()
                        titulo_obj.categorias.add(otros)

                messages.success(request, f"Título «{titulo_obj.titulo}» cargado (stock inicial: 1).")
                return redirect("libros_list")

            except IntegrityError:
                messages.error(request, "Ya existe un título con ese nombre.")
                return render(request, "libro_form.html", {
                    "form": form,
                    "categorias_all": Categoria.objects.order_by("nombre"),
                })

        # form inválido → re-render con errores
        return render(request, "libro_form.html", {
            "form": form,
            "categorias_all": Categoria.objects.order_by("nombre"),
        })
    

  

    # GET
    form = LibroForm()
    return render(request, "libro_form.html", {
        "form": form,
        "categorias_all": Categoria.objects.order_by("nombre"),
    })

  #LIBRO DELETE 
@require_POST
def libro_eliminar(request,pk):
        libro = get_object_or_404(Titulo, pk=pk)
        libro.delete()
        messages.success(request, "Libro eliminado correctamente")
        return redirect('libros_list')

@require_POST
def categoria_create_ajax(request):
    """
    Crea una categoría por AJAX.
    Respuesta JSON: { ok: true, id, nombre, created, message }
    """
    nombre = (request.POST.get("nombre") or "").strip()
    if not nombre:
        return JsonResponse({"ok": False, "error": "El nombre es obligatorio."}, status=400)

    cat, created = Categoria.objects.get_or_create(nombre=nombre)
    msg = f"Categoría «{cat.nombre}» {'creada' if created else 'ya existente, seleccionada'}."
    return JsonResponse({"ok": True, "id": cat.id, "nombre": cat.nombre, "created": created, "message": msg})


@require_POST
@transaction.atomic
def categoria_delete(request, pk):
    """
    Elimina una categoría reasignando sus títulos a 'Otros' si quedarían sin categorías.
    """
    cat = get_object_or_404(Categoria, pk=pk)
    if cat.nombre == "Otros":
        messages.warning(request, "La categoría 'Otros' no se puede eliminar.")
        return redirect(request.META.get("HTTP_REFERER", "libros_list"))

    otros = ensure_categoria_otros()

    titulos = Titulo.objects.filter(categorias=cat).distinct()
    for t in titulos:
        t.categorias.remove(cat)
        if t.categorias.count() == 0:
            t.categorias.add(otros)

    cat.delete()
    messages.success(request, "Categoría eliminada. Los títulos quedaron reasignados a 'Otros' cuando correspondía.")
    return redirect(request.META.get("HTTP_REFERER", "libros_list"))


# -------------------------------------------------------------------
# Préstamos: listado / crear / renovar / devolver / eliminar
# -------------------------------------------------------------------
def prestamos_list(request):
    data = list(
        Prestamo.objects
        .select_related("ejemplar__titulo")
        .values("id", "alumno_nombre", "alumno_dni", "ejemplar__titulo__titulo", "fecha_prestamo", "vence", "estado")
        .order_by("-fecha_prestamo")
    )
    for d in data:
        d["alumno"] = d.pop("alumno_nombre")
        d["dni"] = d.pop("alumno_dni")
        d["libro"] = d.pop("ejemplar__titulo__titulo")
    return render(request, "prestamos_list.html", {"prestamos": data})


def prestamo_create(request):
    hoy = date.today()
    max_date = hoy + timedelta(days=30)

    libros_opts = list(
        Titulo.objects.filter(ejemplares__estado="DISPONIBLE")
        .distinct()
        .values_list("titulo", flat=True)
        .order_by("titulo")
    )

    if request.method == "POST":
        alumno = (request.POST.get("alumno") or request.POST.get("nombre") or "").strip()
        dni    = (request.POST.get("dni") or "").strip()
        titulo_elegido = (request.POST.get("libro") or "").strip()
        f_str  = (request.POST.get("fecha_devolucion") or request.POST.get("fecha_vencimiento") or "").strip()

        try:
            vence_dt = datetime.strptime(f_str, "%Y-%m-%d").date()
        except ValueError:
            return render(request, "prestamo_form.html", {
                "error": "Fecha inválida.",
                "libros_opts": libros_opts, "hoy": hoy, "max_date": max_date,
                "nombre": alumno, "dni": dni, "libro": titulo_elegido,
                "fecha_vencimiento": f_str,
            })

        if not (hoy <= vence_dt <= max_date):
            return render(request, "prestamo_form.html", {
                "error": "La fecha debe estar entre hoy y 30 días.",
                "libros_opts": libros_opts, "hoy": hoy, "max_date": max_date,
                "nombre": alumno, "dni": dni, "libro": titulo_elegido,
                "fecha_vencimiento": f_str,
            })

        if vence_dt.weekday() in (5, 6):
            return render(request, "prestamo_form.html", {
                "error": "La fecha de devolución no puede ser sábado ni domingo.",
                "libros_opts": libros_opts, "hoy": hoy, "max_date": max_date,
                "nombre": alumno, "dni": dni, "libro": titulo_elegido,
                "fecha_vencimiento": f_str,
            })

        ej = (
            Ejemplar.objects
            .select_related("titulo")
            .filter(titulo__titulo=titulo_elegido, estado="DISPONIBLE")
            .first()
        )
        if not ej:
            return render(request, "prestamo_form.html", {
                "error": "El título no tiene ejemplares disponibles.",
                "libros_opts": libros_opts, "hoy": hoy, "max_date": max_date,
                "nombre": alumno, "dni": dni, "libro": titulo_elegido,
                "fecha_vencimiento": f_str,
            })

        with transaction.atomic():
            prestamo = Prestamo.objects.create(
                ejemplar=ej,
                alumno_nombre=alumno,
                alumno_dni=dni,
                fecha_prestamo=hoy,
                vence=vence_dt,
                estado="ACTIVO",
            )
            ej.estado = "PRESTADO"
            ej.save(update_fields=["estado"])

        messages.success(request, f"Préstamo registrado para {alumno} · {ej.titulo.titulo}.")
        return redirect("prestamos_list")

    return render(request, "prestamo_form.html", {
        "libros_opts": libros_opts, "hoy": hoy, "max_date": max_date
    })


def _mover_a_lunes_si_findes(d: date) -> date:
    """Si cae sábado (5) o domingo (6), corre hasta el lunes siguiente."""
    while d.weekday() in (5, 6):
        d += timedelta(days=1)
    return d


@require_POST
def prestamo_renovar(request, pk):
    """
    Suma 7 días al vencimiento si el préstamo está ACTIVO o RENOVADO.
    No permite renovar si está DEVUELTO.
    """
    prestamo = get_object_or_404(Prestamo.objects.select_related("ejemplar__titulo"), pk=pk)

    if prestamo.estado == "DEVUELTO":
        messages.warning(request, "No podés renovar un préstamo devuelto.")
        return redirect("prestamos_list")

    # Si quisieras bloquear renovar vencidos, descomentá:
    # if prestamo.vence < date.today():
    #     messages.warning(request, "El préstamo está vencido: primero debe regularizarse.")
    #     return redirect("prestamos_list")

    prestamo.vence = prestamo.vence + timedelta(days=7)
    prestamo.estado = "RENOVADO"
    prestamo.save(update_fields=["vence", "estado"])
    messages.success(request, f"Préstamo renovado +7 días para «{prestamo.ejemplar.titulo.titulo}». Nuevo vencimiento: {prestamo.vence}.")
    return redirect("prestamos_list")


@require_POST
def prestamo_devolver(request, pk):
    """
    Marca el préstamo como DEVUELTO y libera el ejemplar (DISPONIBLE).
    """
    prestamo = get_object_or_404(Prestamo.objects.select_related("ejemplar__titulo"), pk=pk)

    if prestamo.estado == "DEVUELTO":
        messages.info(request, "Ese préstamo ya estaba marcado como devuelto.")
        return redirect("prestamos_list")

    with transaction.atomic():
        prestamo.estado = "DEVUELTO"
        prestamo.save(update_fields=["estado"])
        ej = prestamo.ejemplar
        if ej and ej.estado != "DISPONIBLE":
            ej.estado = "DISPONIBLE"
            ej.save(update_fields=["estado"])

    messages.success(request, f"Se marcó como devuelto: {prestamo.ejemplar.titulo.titulo}.")
    return redirect("prestamos_list")


@require_POST
def prestamo_eliminar(request, pk):
    """
    Elimina el préstamo. Si no estaba devuelto, primero libera el ejemplar.
    """
    prestamo = get_object_or_404(Prestamo.objects.select_related("ejemplar"), pk=pk)

    with transaction.atomic():
        if prestamo.estado != "DEVUELTO" and prestamo.ejemplar:
            ej = prestamo.ejemplar
            if ej.estado != "DISPONIBLE":
                ej.estado = "DISPONIBLE"
                ej.save(update_fields=["estado"])

        prestamo.delete()

    messages.success(request, "Préstamo eliminado.")
    return redirect("prestamos_list")
